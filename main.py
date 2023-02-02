#!/usr/bin/env python
# -*- coding: utf8 -*-
import sys
import tkinter as tk
import webbrowser
from functools import partial
from nfc_reader import get_idm, get_stunum
from datetime import datetime
import gspread
import json
from oauth2client.service_account import ServiceAccountCredentials
import openpyxl as px
import random
from openpyxl.styles.borders import Border, Side
flag_quit=0
import requests
import time
from module import delete_old_messages, copy_to_onedrive

class Rams():
    def __init__(self):
        self.root = tk.Tk()
        img = tk.PhotoImage(file='icon.png')
        self.root.tk.call('wm', 'iconphoto', self.root._w, img)
        self.root.title(u'衛星部屋RAMS')
        self.root.geometry("600x370")
        self.wratio = int(self.root.winfo_screenwidth()/600*0.9)
        self.hratio = int(self.root.winfo_screenheight()/370*0.8)
        self.root.resizable(width=False, height=False)
        self.root.attributes("-fullscreen", True)
        self.username = 'default'
        self.stunum = 'default'
        self.fin_comment = 'default_comment'
        self.error_flag = False
        self.finish()
        self.register_init()
        self.enter_init()
        self.exit_init()
        self.top_init()
        self.root.mainloop()

    def top_init(self):
        self.f_top = tk.Frame(self.root)
        self.root.grid_columnconfigure(0,weight=1)
        self.label = tk.Label(self.f_top, text = u'\n衛星部屋RAMS', font=('', 30*self.wratio))
        self.credit = tk.Label(self.f_top, text = u'Developed by forestwaterfall', font=('', 10*self.wratio))

        self.enter_b = tk.Button(self.f_top, text=u'入室', font=('', 30*self.wratio), width=10*self.wratio, height=2*self.hratio, bg='bisque', relief='raised', bd=2, command=self.enter_init)
        self.enter_b.grid(row=2, column=0, padx=10*self.wratio, pady=40*self.hratio)
        self.exit_b = tk.Button(self.f_top, text=u'退室', font=('', 30*self.wratio), width=10*self.wratio, height=2*self.hratio, bg='bisque', relief='raised', bd=2, command=self.exit_init)
        self.exit_b.grid(row=2, column=1, padx=10*self.wratio, pady=40*self.hratio)
        self.label.grid(row=0, columnspan=2, pady=10*self.hratio)
        self.credit.grid(row=1, columnspan=2, pady=0*self.hratio)
        self.f_top.grid_columnconfigure(1,weight=1)
        self.f_top.grid_columnconfigure(0,weight=1)
        self.register_b = tk.Button(self.f_top, text=u'新規登録', font=('', 10*self.wratio), width=20*self.wratio, height=2*self.hratio, bg='bisque', relief='raised', bd=2, command=self.register_init)
        self.register_b.grid(row=4, column=0, columnspan=2, padx=10*self.wratio, pady=10*self.hratio)
        self.exit_rams = tk.Button(self.f_top, text=u'終了', font=('', 10*self.wratio), width=20*self.wratio, height=2*self.hratio, bg='bisque', relief='raised', bd=2, command=self.end_rams)
        self.exit_rams.grid(row=5, column=0, columnspan=2, padx=10*self.wratio, pady=10*self.hratio)

        self.f_top.grid(row=0, column=0, sticky='nsew')

    def end_rams(self):
        sys.exit()

    def openhistory(self):
        webbrowser.open('https://docs.google.com/spreadsheets/d/15ynrekrU8WPaEWGmF2BCN5Eu_QReAJPF8P0QHnamdOs/edit?usp=sharing')

    def backtop(self):
        self.error_flag = False
        self.top_init()
        self.f_top.tkraise()

    def enter_init(self):
        self.f_enter = tk.Frame(self.root)
        self.f_enter.grid(row=0, column=0, sticky='nsew')
        self.root.grid_columnconfigure(0,weight=1)

        self.backtop_b = tk.Button(self.f_enter, text=u'TOP画面に戻る', font=('', 8*self.wratio), width=15*self.wratio, height=2*self.hratio, bg='bisque', relief='raised', bd=2, command=self.backtop)
        self.backtop_b.grid(row=0, column=0, padx=10*self.wratio, pady=10*self.hratio)
        self.f_enter.grid_columnconfigure(0,weight=0)
        self.manual_b = tk.Button(self.f_enter, text=u'手入力', font=('', 8*self.wratio), width=15*self.wratio, height=2*self.hratio, bg='bisque', relief='raised', bd=2, command=self.enter_manu)
        self.manual_b.grid(row=1, column=0, padx=10*self.wratio, pady=10*self.hratio)
        self.title_l = tk.Label(self.f_enter, text='\n\n入室', font=('',25*self.wratio))
        self.title_l.grid(row=0, column=3, sticky='w')
        self.time_l = tk.Label(self.f_enter, anchor='e', text='入室時刻：', font=('',10*self.wratio))
        self.time_l.grid(row=1, column=2, pady=10*self.hratio, sticky='e')
        self.time_e = tk.Entry(self.f_enter, width=10*self.wratio, font=('',20*self.hratio))
        self.time_now = datetime.now().strftime('%m/%d %H:%M')
        self.time_e.insert(0, self.time_now)
        self.time_e.grid(row=1, column=3, sticky='w')
        self.thermo_l = tk.Label(self.f_enter, anchor='e', text='体温：', font=('',10*self.wratio))
        self.thermo_l.grid(row=2, column=2, pady=10*self.hratio, sticky='e')
        self.thermo_e = tk.Entry(self.f_enter, width=10*self.wratio, font=('',20*self.hratio))
        self.thermo_e.grid(row=2, column=3, sticky='w')
        self.thermo_e.focus_set()
        self.pleasetouch_l = tk.Label(self.f_enter, text='\n\n体温を入力後，学生証をカードリーダーにタッチして\n送信を押してください．', font=('', 15*self.wratio))
        self.pleasetouch_l.grid(row=4, column=2, columnspan=2, pady=20*self.hratio, sticky='e')
        self.submitenter_b = tk.Button(self.f_enter, text=u'送信', font=('', 15*self.wratio), width=15*self.wratio, height=2*self.hratio, bg='bisque', relief='raised', bd=2, command=self.submit_enter)
        self.submitenter_b.grid(row=5, column=2, columnspan=2, padx=10*self.wratio, pady=5*self.hratio)

        self.f_enter.tkraise()

    def submit_enter(self):
        self.idm = get_idm()
        self.idm_check()
        if str(self.idm)=="":
            self.pleasetouch_l = tk.Label(self.f_enter, text='学生証が認識できません．', fg='red', font=('', 15*self.wratio))
            self.pleasetouch_l.grid(row=6, column=1, columnspan=3, pady=5*self.hratio)
        elif self.thermo_e.get() == "":
            self.pleasetouch_l = tk.Label(self.f_enter, text='体温が入力されていません．', fg='red', font=('', 15*self.wratio))
            self.pleasetouch_l.grid(row=6, column=1, columnspan=3, pady=5*self.hratio)
        else:
            self.idm_check()
            if self.error_flag:
                self.finish()
                self.f_hello.tkraise()
                self.root.after(3000,self.backtop)
            else:
                self.fin_comment = 'ようこそ，'+self.username+'さん．'
                self.finish()
                self.f_hello.tkraise()
                self.root.after(1000,self.idm_enter)

    def idm_enter(self):
        book = px.load_workbook('data.xlsx')
        ws = book.get_sheet_by_name('idm_stunum')
        i = 1
        print(str(self.idm),"のidmで入室しようとしています")
        while(True):
            if str(self.idm) == str(ws['A%d'%i].value):
                self.stunum = ws['B%d'%i].value
                book.save('data.xlsx')
                self.stunum_enter()
                break
            elif str(ws['A%d'%i].value) == 'None':
                self.error_flag = True
                break
            i = i+1

    def stunum_enter(self):
        book = px.load_workbook('data.xlsx')
        ws = book.get_sheet_by_name('user_list')
        i = 1
        while(True):
            if str(self.stunum) == str(ws['A%d'%i].value):
                self.username = ws['B%d'%i].value
                ws['C%d'%i].value = 'in'

                book2 = px.load_workbook('history.xlsx')
                worksheet = book2.get_sheet_by_name('history')
                cell_to_put = 1
                while str(worksheet['A%d'%cell_to_put].value) != 'None':
                    cell_to_put += 1
                worksheet.cell(cell_to_put, 1).value = self.username
                worksheet.cell(cell_to_put, 2).value = self.time_e.get()
                worksheet.cell(cell_to_put, 4).value = self.thermo_e.get()
                book2.save('history.xlsx')

                ws['D%i'%i].value = cell_to_put
                ws['E%d'%i].value = str(time.time())
                book.save('data.xlsx')
                self.post_slack(self.username, 'in')
                copy_to_onedrive()
                self.backtop()
                break
            elif str(ws['A%d'%i].value) == 'None':
                self.error_flag = True
                break
            i = i+1



    def exit_init(self):
        self.f_exit = tk.Frame(self.root)
        self.f_exit.grid(row=0, column=0, sticky='nsew')
        self.root.grid_columnconfigure(0,weight=1)

        self.backtop_b = tk.Button(self.f_exit, text=u'TOP画面に戻る', font=('', 8*self.wratio), width=15*self.wratio, height=2*self.hratio, bg='bisque', relief='raised', bd=2, command=self.backtop)
        self.backtop_b.grid(row=0, column=0, padx=15*self.wratio, pady=10*self.hratio)
        self.f_exit.grid_columnconfigure(0,weight=0)
        self.manual_b = tk.Button(self.f_exit, text=u'手入力', font=('', 8*self.wratio), width=15*self.wratio, height=2*self.hratio, bg='bisque', relief='raised', bd=2, command=self.exit_manu)
        self.manual_b.grid(row=1, column=0, padx=10*self.wratio, pady=10*self.hratio)
        self.title_l = tk.Label(self.f_exit, text='\n\n退室', font=('',25*self.wratio))
        self.title_l.grid(row=0, column=3, sticky='w')
        self.time_l = tk.Label(self.f_exit, anchor='e', text='退室時刻：', font=('',10*self.wratio))
        self.time_l.grid(row=1, column=2, pady=10*self.hratio, sticky='e')
        self.time_e = tk.Entry(self.f_exit, width=10*self.wratio, font=('',20*self.hratio))
        self.time_now = datetime.now().strftime('%m/%d %H:%M')
        self.time_e.insert(0, self.time_now)
        self.time_e.grid(row=1, column=3, sticky='w')
        self.pleasetouch_l = tk.Label(self.f_exit, text='\n学生証をカードリーダーにタッチして\n送信を押してください．', font=('', 15*self.wratio))
        self.pleasetouch_l.grid(row=3, column=2, columnspan=2, pady=20*self.hratio, sticky='e')
        self.submitenter_b = tk.Button(self.f_exit, text=u'送信', font=('', 15*self.wratio), width=15*self.wratio, height=2*self.hratio, bg='bisque', relief='raised', bd=2, command=self.submit_exit)
        self.submitenter_b.grid(row=4, column=2, columnspan=2, padx=10*self.wratio, pady=5*self.hratio)
        self.f_exit.tkraise()

    def submit_exit(self):
        self.idm = get_idm()
        self.idm_check()
        if str(self.idm)=="":
            self.pleasetouch_l = tk.Label(self.f_exit, text='学生証が認識できません．', fg='red', font=('', 15*self.wratio))
            self.pleasetouch_l.grid(row=5, column=1, columnspan=3, pady=5*self.hratio)
        else:
            self.idm_check()
            if self.error_flag:
                self.finish()
                self.f_hello.tkraise()
                self.root.after(3000,self.backtop)
            else:
                self.fin_comment = 'さようなら，'+self.username+'さん．'
                self.finish()
            self.f_hello.tkraise()
            self.root.after(1000,self.idm_exit)

    def idm_exit(self):
        book = px.load_workbook('data.xlsx')
        ws = book.get_sheet_by_name('idm_stunum')
        i = 1
        print(str(self.idm)+"のidmで退室しようとしています")
        while(True):
            if str(self.idm) == str(ws['A%d'%i].value):
                self.stunum = ws['B%d'%i].value
                book.save('data.xlsx')
                self.stunum_exit()
                break
            elif str(ws['A%d'%i].value) == 'None':
                self.error_flag = True
                break
            i = i+1

    def stunum_exit(self):
        book = px.load_workbook('data.xlsx')
        ws = book.get_sheet_by_name('user_list')
        i = 1
        while(True):
            if str(self.stunum) == str(ws['A%d'%i].value):
                self.username = ws['B%d'%i].value
                book2 = px.load_workbook('history.xlsx')
                worksheet = book2.get_sheet_by_name('history')
                if ws['C%d'%i].value == 'in':
                    ws['C%d'%i].value = 'out'


                    cell_to_put = 1
                    while str(worksheet['A%d'%cell_to_put].value )!= 'None':
                        cell_to_put += 1
                    if str(worksheet['C%d'%cell_to_put].value) != 'None':
                        worksheet.cell(cell_to_put, 1).value = self.username
                        worksheet.cell(cell_to_put, 3).value = self.time_e.get()
                    else:
                        cell_to_put = ws['D%d'%i].value
                        worksheet.cell(cell_to_put, 1).value = self.username
                        worksheet.cell(cell_to_put, 3).value = self.time_e.get()
                else:
                    ws['C%d'%i].value = 'out'
                    cell_to_put = 1
                    while str(worksheet['A%d'%cell_to_put].value) != 'None':
                        cell_to_put += 1
                    worksheet.cell(cell_to_put, 1).value = self.username
                    worksheet.cell(cell_to_put, 3).value = self.time_e.get()
                ws['D%i'%i].value = cell_to_put
                book.save('data.xlsx')
                book2.save('history.xlsx')
                self.post_slack(self.username, 'out')
                copy_to_onedrive()
                self.backtop()
                break
            elif str(ws['A%d'%i].value) == 'None':
                self.error_flag = True
                break
            i = i+1

    def post_slack(self, username, move):
        book = px.load_workbook('data.xlsx')
        ws = book.get_sheet_by_name('user_list')
        i = 1
        in_list = []
        while str(ws['A%d'%i].value) != 'None':
            i += 1
            if ws['C%d'%i].value == 'in':
                enter_time = float(ws['E%d'%i].value)
                if time.time() - enter_time > 172800:
                    ws['C%d'%i].value = 'out'
                    remove_user = ws['B%d'%i].value
                    book.save('data.xlsx')
                    self.remove(remove_user)
                    continue
                in_list.append(ws['B%d'%i].value)
        num_in = len(in_list)
        if move == 'in':
            msg = username + 'さんが入室しました．\n'
            icon = ':inred:'
            postname = '衛星部屋入退室管理システム RAMS(入室)'
        elif move == 'out':
            msg = username + 'さんが退室しました．\n'
            icon = 'outblue'
            postname = '衛星部屋入退室管理システム RAMS(退室)'
        elif move == 'remove':
            msg = username + 'さんは48時間以上在室状態になっているため、自動的に退室扱いとなりました．\n現在在室中の場合は再度入室を行ってください．\n'
            icon = 'outblue'
            postname = '衛星部屋入退室管理システム RAMS(退室)'
        if len(in_list) == 0:
            msg += '現在衛星部屋には誰もいません．'
        else:
            msg += '現在 *' + str(len(in_list)) + '人* が衛星部屋にいます．\n('
            for person in in_list:
                msg += person + '，'
            msg = msg[:-1] + ')'
        msg += "\n---------------------------"
        print(msg+"\n")
        WEB_HOOK_URL_SSSRC = "https://hooks.slack.com/services/XXXXX"
        WEB_HOOK_URL_SSSUP = "https://hooks.slack.com/services/XXXXX"
        data = json.dumps({
            # メッセージ内容
            "text" : msg,
            # アイコン
            "icon_emoji" : icon,
            # 投稿者名
            "username" : postname
        })
        try:
            requests.post(WEB_HOOK_URL_SSSRC, data)
            requests.post(WEB_HOOK_URL_SSSUP, data)
            delete_old_messages()
        except:
            pass


    def remove(self, username):
        msg = username + 'さんは48時間以上在室状態になっているため、自動的に退室扱いとなりました．\n現在在室中の場合は再度入室を行ってください．\n'
        icon = 'outblue'
        postname = '衛星部屋入退室管理システム RAMS(退室)'

        print(msg+"\n")
        msg += "\n---------------------------"
        WEB_HOOK_URL_SSSRC = "https://hooks.slack.com/services/XXXXX"
        WEB_HOOK_URL_SSSUP = "https://hooks.slack.com/services/XXXXX"
        data = json.dumps({
            # メッセージ内容
            "text" : msg,
            # アイコン
            "icon_emoji" : icon,
            # 投稿者名
            "username" : postname
        })
        try:
            requests.post(WEB_HOOK_URL_SSSRC, data)
            requests.post(WEB_HOOK_URL_SSSUP, data)
            delete_old_messages()
        except:
            pass


    def register_init(self):
        self.f_register = tk.Frame(self.root)
        self.f_register.grid(row=0, column=0, sticky='nsew')
        self.root.grid_columnconfigure(0,weight=1)

        self.backtop_b = tk.Button(self.f_register, text=u'TOP画面に戻る', font=('', 8*self.wratio), width=15*self.wratio, height=2*self.hratio, bg='bisque', relief='raised', bd=2, command=self.backtop)
        self.backtop_b.grid(row=0, column=0, padx=15*self.wratio, pady=10*self.hratio)
        self.f_register.grid_columnconfigure(0,weight=0)
        self.title_l = tk.Label(self.f_register, text='\n\n新規登録', font=('',25*self.wratio))
        self.title_l.grid(row=0, column=3, sticky='w')
        self.stunum_l = tk.Label(self.f_register, anchor='e', text='学籍番号：', font=('',10*self.wratio))
        self.stunum_l.grid(row=1, column=2, pady=10*self.hratio, sticky='e')
        self.stunum_e = tk.Entry(self.f_register, width=30, textvariable=self.stunum)
        self.stunum_e.insert(0, '')
        self.stunum_e.grid(row=1, column=3, sticky='w')
        self.name_l = tk.Label(self.f_register, anchor='e', text='名前：', font=('',10*self.wratio))
        self.name_l.grid(row=2, column=2, pady=10*self.hratio, sticky='e')
        self.name_e = tk.Entry(self.f_register, width=10*self.wratio, font=('',20*self.hratio))
        self.name_e.grid(row=2, column=3, sticky='w')
        self.name_e.focus_set()
        self.pleasetouch_l = tk.Label(self.f_register, text='\n学生証をカードリーダーにタッチして\n送信を押してください．', font=('', 15*self.wratio))
        self.pleasetouch_l.grid(row=3, column=2, columnspan=2, pady=20*self.hratio, sticky='e')
        self.overwrite_l = tk.Label(self.f_register, text='(学生証が既に登録されている場合，情報が上書きされます．)', font=('', 10*self.wratio))
        self.overwrite_l.grid(row=4, column=2, columnspan=2, pady=10*self.hratio, sticky='e')
        self.submitenter_b = tk.Button(self.f_register, text=u'送信', font=('', 15*self.wratio), width=15*self.wratio, height=2*self.hratio, bg='bisque', relief='raised', bd=2, command=lambda: self.submit_register(self.stunum_e.get()))
        self.submitenter_b.grid(row=5, column=2, columnspan=2, padx=10*self.wratio, pady=5*self.hratio)
        self.f_register.tkraise()

    def submit_register(self, stunum):
        self.stunum = stunum
        self.idm = get_idm()
        if str(self.idm)=="":
            self.pleasetouch_l = tk.Label(self.f_register, text='学生証が認識できません．', fg='red', font=('', 15*self.wratio))
            self.pleasetouch_l.grid(row=6, column=1, columnspan=3, pady=5*self.hratio)
        else:
            self.stunum = self.stunum_e.get()
            self.username = self.name_e.get()
            if str(self.stunum) != '' and str(self.username) != '':
                self.fin_comment = '登録が完了しました．'
                self.register_fin()
                self.finish()
                self.f_hello.tkraise()
            else:
                self.error_fin()
                self.f_bye.tkraise()
            self.root.after(3000,self.backtop)

    def finish(self):
        self.f_hello = tk.Frame(self.root)
        self.f_hello.grid(row=0, column=0, sticky='nsew')
        self.root.grid_columnconfigure(0,weight=1)

        if(self.error_flag):
            self.hello_l = tk.Label(self.f_hello, text='ERROR：登録データがありません', fg='red', font=('',25*self.wratio))
        else:
            self.hello_l = tk.Label(self.f_hello, text=self.fin_comment, font=('',25*self.wratio))
        self.hello_l.grid(row=0,column=0,pady=20*self.hratio)
        self.callback_l = tk.Label(self.f_hello, text='自動でトップ画面に戻ります', font=('',15*self.wratio))
        self.callback_l.grid(row=2,column=0,pady=20*self.hratio)
        self.f_hello.grid_columnconfigure(0,weight=1)
        self.f_hello.grid_rowconfigure(0,weight=1)
        self.f_hello.tkraise()


    def register_fin(self):
        book = px.load_workbook('data.xlsx')
        ws = book.get_sheet_by_name('idm_stunum')
        cell_to_put = 1
        while(True):
            if str(ws['A%d'%cell_to_put].value) == str(self.idm):
                ws['B%d'%cell_to_put].value = self.stunum
                break
            elif str(ws['A%d'%cell_to_put].value) == 'None':
                ws['A%d'%cell_to_put].value = str(self.idm)
                ws['B%d'%cell_to_put].value = self.stunum
                break
            cell_to_put = cell_to_put + 1
        ws = book.get_sheet_by_name('user_list')
        cell_to_put = 1
        while(True):
            if str(ws['A%d'%cell_to_put].value) == self.stunum:
                ws['B%d'%cell_to_put].value = self.username
                break
            elif str(ws['A%d'%cell_to_put].value) == 'None':
                ws['A%d'%cell_to_put].value = self.stunum
                ws['B%d'%cell_to_put].value = self.username
                break
            cell_to_put = cell_to_put + 1
        book.save('data.xlsx')


    def enter_manu(self):
        self.f_enter = tk.Frame(self.root)
        self.f_enter.grid(row=0, column=0, sticky='nsew')
        self.root.grid_columnconfigure(0,weight=1)

        self.backtop_b = tk.Button(self.f_enter, text=u'TOP画面に戻る', font=('', 8*self.wratio), width=15*self.wratio, height=2*self.hratio, bg='bisque', relief='raised', bd=2, command=self.backtop)
        self.backtop_b.grid(row=0, column=0, padx=10*self.wratio, pady=10*self.hratio)
        self.f_enter.grid_columnconfigure(0,weight=0)
        self.title_l = tk.Label(self.f_enter, text='\n\n入室', font=('',25*self.wratio))
        self.title_l.grid(row=0, column=3, sticky='w')
        self.time_l = tk.Label(self.f_enter, anchor='e', text='入室時刻：', font=('',10*self.wratio))
        self.time_l.grid(row=1, column=2, pady=10*self.hratio, sticky='e')
        self.time_e = tk.Entry(self.f_enter, width=10*self.wratio, font=('',20*self.hratio))
        self.time_now = datetime.now().strftime('%m/%d %H:%M')
        self.time_e.insert(0, self.time_now)
        self.time_e.grid(row=1, column=3, sticky='w')
        self.thermo_l = tk.Label(self.f_enter, anchor='e', text='体温：', font=('',10*self.wratio))
        self.thermo_l.grid(row=2, column=2, pady=10*self.hratio, sticky='e')
        self.thermo_e = tk.Entry(self.f_enter, width=10*self.wratio, font=('',20*self.hratio))
        self.thermo_e.grid(row=2, column=3, sticky='w')
        self.thermo_e.focus_set()
        self.stunum_l = tk.Label(self.f_enter, anchor='e', text='学籍番号：', font=('',10*self.wratio))
        self.stunum_l.grid(row=3, column=2, pady=10*self.hratio, sticky='e')
        self.stunum_e = tk.Entry(self.f_enter, width=10*self.wratio, font=('',20*self.hratio))
        self.stunum_e.grid(row=3, column=3, sticky='w')
        self.pleasetouch_l = tk.Label(self.f_enter, text='\n体温・学籍番号を入力後，送信を押してください．', font=('', 15*self.wratio))
        self.pleasetouch_l.grid(row=4, column=2, columnspan=2, pady=20*self.hratio, sticky='e')
        self.submitenter_b = tk.Button(self.f_enter, text=u'送信', font=('', 15*self.wratio), width=15*self.wratio, height=2*self.hratio, bg='bisque', relief='raised', bd=2, command=self.entermanu_fin)
        self.submitenter_b.grid(row=5, column=2, columnspan=2, padx=10*self.wratio, pady=5*self.hratio)

        self.f_enter.tkraise()

    def entermanu_fin(self):
        if self.thermo_e.get() == "":
            self.pleasetouch_l = tk.Label(self.f_enter, text='体温が入力されていません．', fg='red', font=('', 15*self.wratio))
            self.pleasetouch_l.grid(row=4, column=1, columnspan=3, pady=5*self.hratio)
            return 0

        self.fin_comment = 'ようこそ，'+self.username+'さん．'
        self.stunum = self.stunum_e.get()
        self.stunum_check()
        if self.error_flag:
            self.finish()
            self.root.after(3000,self.backtop)
        else:
            self.stunum_check()
            if self.error_flag:
                self.finish()
                self.f_hello.tkraise()
                self.root.after(3000,self.backtop)
            else:
                self.fin_comment = 'ようこそ，'+self.username+'さん．'
                self.finish()
                self.f_hello.tkraise()
                self.root.after(1000,self.stunum_enter)



    def exit_manu(self):
        self.f_exit = tk.Frame(self.root)
        self.f_exit.grid(row=0, column=0, sticky='nsew')
        self.root.grid_columnconfigure(0,weight=1)

        self.backtop_b = tk.Button(self.f_exit, text=u'TOP画面に戻る', font=('', 8*self.wratio), width=15*self.wratio, height=2*self.hratio, bg='bisque', relief='raised', bd=2, command=self.backtop)
        self.backtop_b.grid(row=0, column=0, padx=15*self.wratio, pady=10*self.hratio)
        self.f_exit.grid_columnconfigure(0,weight=0)
        self.title_l = tk.Label(self.f_exit, text='\n\n退室', font=('',25*self.wratio))
        self.title_l.grid(row=0, column=3, sticky='w')
        self.time_l = tk.Label(self.f_exit, anchor='e', text='退室時刻：', font=('',10*self.wratio))
        self.time_l.grid(row=1, column=2, pady=10*self.hratio, sticky='e')
        self.time_e = tk.Entry(self.f_exit, width=10*self.wratio, font=('',20*self.hratio))
        self.time_now = datetime.now().strftime('%m/%d %H:%M')
        self.time_e.insert(0, self.time_now)
        self.time_e.grid(row=1, column=3, sticky='w')
        self.stunum_l = tk.Label(self.f_exit, anchor='e', text='学籍番号：', font=('',10*self.wratio))
        self.stunum_l.grid(row=2, column=2, pady=10*self.hratio, sticky='e')
        self.stunum_e = tk.Entry(self.f_exit, width=10*self.wratio, font=('',20*self.hratio))
        self.stunum_e.grid(row=2, column=3, sticky='w')
        self.stunum_e.focus_set()
        self.pleasetouch_l = tk.Label(self.f_exit, text='\n学籍番号を入力後，送信を押してください．', font=('', 15*self.wratio))
        self.pleasetouch_l.grid(row=3, column=2, columnspan=2, pady=20*self.hratio, sticky='e')
        self.submitenter_b = tk.Button(self.f_exit, text=u'送信', font=('', 15*self.wratio), width=15*self.wratio, height=2*self.hratio, bg='bisque', relief='raised', bd=2, command=self.exitmanu_fin)
        self.submitenter_b.grid(row=4, column=2, columnspan=2, padx=10*self.wratio, pady=5*self.hratio)
        self.f_exit.tkraise()

    def exitmanu_fin(self):
        self.fin_comment = 'さようなら，'+self.username+'さん．'
        self.stunum = self.stunum_e.get()
        self.stunum_check()
        if self.error_flag:
            self.finish()
            self.root.after(3000,self.backtop)
        else:
            self.stunum_check()
            if self.error_flag:
                self.finish()
                self.f_hello.tkraise()
                self.root.after(3000,self.backtop)
            else:
                self.fin_comment = 'さようなら，'+self.username+'さん．'
                self.finish()
                self.f_hello.tkraise()
                self.root.after(1000,self.stunum_exit)

    def error_fin(self):
        self.f_bye = tk.Frame(self.root)
        self.f_bye.grid(row=0, column=0, sticky='nsew')
        self.root.grid_columnconfigure(0,weight=1)

        self.hello_l = tk.Label(self.f_bye, text='入力が不足しているか，正しくありません．', fg='red', font=('',25*self.wratio))
        self.hello_l.grid(row=0,column=0,pady=20*self.hratio)
        self.count = 5
        self.callback_l = tk.Label(self.f_bye, text='自動でトップ画面に戻ります', font=('',15*self.wratio))
        self.callback_l.grid(row=2,column=0,pady=20*self.hratio)
        self.f_bye.grid_columnconfigure(0,weight=1)
        self.f_bye.grid_rowconfigure(0,weight=1)
        self.f_bye.tkraise()
        self.root.after(3000,self.backtop)

    def idm_check(self):
        book = px.load_workbook('data.xlsx')
        ws = book['idm_stunum']
        i = 1
        while(True):
            if str(self.idm) == str(ws['A%d'%i].value):
                self.stunum = ws['B%d'%i].value
                self.stunum_check()
                break
            elif str(ws['A%d'%i].value) == 'None':
                only_stunum_flag = self.check_only_stunum()
                if not only_stunum_flag:
                    self.error_flag = True
                break
            i = i+1

    def check_only_stunum(self):
        self.stunum = get_stunum()
        if self.stunum == '':
            return False
        elif self.stunum_check() == False:
            return False
        else:
            book = px.load_workbook('data.xlsx')
            ws = book.get_sheet_by_name('idm_stunum')
            cell_to_put = 1
            while(True):
                if str(ws['A%d'%cell_to_put].value) == str(self.idm):
                    ws['B%d'%cell_to_put].value = self.stunum
                    break
                elif str(ws['A%d'%cell_to_put].value) == 'None':
                    ws['A%d'%cell_to_put].value = str(self.idm)
                    ws['B%d'%cell_to_put].value = self.stunum
                    break
                cell_to_put = cell_to_put + 1
            book.save('data.xlsx')
            return True

    def stunum_check(self):
        book = px.load_workbook('data.xlsx')
        ws = book['user_list']
        i = 1
        while(True):
            if str(self.stunum) == str(ws['A%d'%i].value):
                self.username = ws['B%d'%i].value
                return True
            elif str(ws['A%d'%i].value) == 'None':
                self.error_flag = True
                return False
            i = i+1

if __name__ == '__main__':
    print("RAMS running...")
    f=Rams()
